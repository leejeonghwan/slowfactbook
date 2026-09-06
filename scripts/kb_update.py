#!/usr/bin/env python3
"""
KB부동산 데이터허브 월간 엑셀로 기존 차트를 제자리 갱신한다.

KB 자료는 KOSIS에 없다. 강남11개구·강북14개구는 KB 고유 구분이고, KOSIS(한국부동산원)에
있는 건 정의가 다른 '실거래 평균가격(만원/㎡)'이라 갖다 붙이면 조용히 틀린 차트가 된다.
그래서 KB는 별도 트랙으로 둔다. 자동 조회 창구가 없으니 엑셀을 받아서 넣는다.

  https://data.kbland.kr/kbstats/wmh?tIdx=HT06&tsIdx=aptSaleAvgPrice → 엑셀 내려받기

등록부는 data/kb_charts.json 이다.
  {
   "c0718": {
    "sheet": "아파트 매매평균가격",
    "series": {"강남11개구": "강남11개구", ...},   차트 계열명 → 엑셀 지역명
    "scale": 0.0001,                              엑셀 만원 → 차트 억 원
    "source": "KB부동산 데이터허브 「월간 아파트 매매평균가격」"
   }
  }

안전장치는 refresh_charts.py 와 같다.
  1. 겹치는 구간을 전부 대조해서 하나라도 어긋나면 그 차트는 건드리지 않는다.
  2. 겹치는 구간의 값은 고치지 않고 뒤에만 이어붙인다.
  3. 계열이 하나라도 엑셀에 없으면 건드리지 않는다.

  python3 scripts/kb_update.py <엑셀경로>            # 미리보기
  python3 scripts/kb_update.py <엑셀경로> --apply
"""
import os, sys, json, re, datetime, argparse

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
sys.path.insert(0, os.path.join(ROOT, "scripts"))
import generate_site as g  # noqa: E402

TOL = 0.005     # 겹치는 구간 허용 오차 0.5%


def read_sheet(path, sheet):
    """엑셀 한 장을 {지역명: [(YYYYMM, 값)]} 으로 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"시트 '{sheet}' 가 없습니다. 있는 시트: {wb.sheetnames}")
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    months = []
    for h in hdr[1:]:
        if hasattr(h, "year"):
            months.append(f"{h.year}{h.month:02d}")
        else:                                   # '2026.08' 같은 문자열도 받아준다
            m = re.match(r"(\d{4})[.\-/ ]?(\d{1,2})", str(h or ""))
            months.append(f"{m.group(1)}{int(m.group(2)):02d}" if m else None)
    out = {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if not nm:
            continue
        vals = []
        for i, c in enumerate(range(2, ws.max_column + 1)):
            v = ws.cell(r, c).value
            vals.append((months[i], v if isinstance(v, (int, float)) else None))
        out[str(nm).strip()] = vals
    return out


def fmt(sample, ym):
    """기존 라벨 생김새대로 새 라벨을 만든다."""
    s, y, mm = str(sample).strip(), ym[:4], int(ym[4:])
    if re.fullmatch(r"(19|20)\d{2}", s):
        return y
    m = re.fullmatch(r"(19|20)\d{2}([-./])\d{1,2}", s)
    if m:
        return f"{y}{m.group(2)}{mm:02d}"
    if re.fullmatch(r"(19|20)\d{2}\s*년", s):
        return f"{y}년"
    return f"{y}-{mm:02d}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args()

    regp = os.path.join(DATA, "kb_charts.json")
    if not os.path.exists(regp):
        print(f"등록부가 없습니다: {regp}")
        return 1
    reg = json.load(open(regp, encoding="utf-8"))

    items = g.load_items(DATA)
    g.assign_ids(items, os.path.join(DATA, "ids.json"))
    by_id = {it["id"]: it for it in items}

    ovp = os.path.join(DATA, "overrides.json")
    ov = json.load(open(ovp, encoding="utf-8")) if os.path.exists(ovp) else {}
    today = datetime.date.today().isoformat()
    done, skip = [], []

    for cid, conf in reg.items():
        it = by_id.get(cid)
        if not it:
            skip.append((cid, "차트를 찾지 못함"))
            continue
        try:
            sheet = read_sheet(a.xlsx, conf["sheet"])
        except Exception as e:
            skip.append((cid, str(e)[:70]))
            continue
        sc = conf.get("scale", 1)
        names = conf.get("series") or {n: n for n in it["seriesNames"]}

        missing = [n for n in it["seriesNames"] if names.get(n, n) not in sheet]
        if missing:
            skip.append((cid, f"엑셀에 없는 계열: {missing}"))
            continue

        # 차트 첫 점이 엑셀 몇 번째 달인지 찾는다 — 값으로 맞춘다
        ref = it["seriesNames"][0]
        col = sheet[names.get(ref, ref)]
        cv = it["series"][0]
        first = next((v for v in cv if v is not None), None)
        off = None
        for j, (ym, xv) in enumerate(col):
            if xv is None or first is None:
                continue
            if abs(first - xv * sc) / max(abs(xv * sc), 1e-9) <= TOL:
                off = j
                break
        if off is None:
            skip.append((cid, "엑셀에서 차트 시작 시점을 못 찾음"))
            continue

        # 겹치는 구간 전수 대조
        bad = 0
        for nm, s in zip(it["seriesNames"], it["series"]):
            xs = sheet[names.get(nm, nm)]
            for i, c in enumerate(s):
                j = i + off
                if c is None or j >= len(xs) or xs[j][1] is None:
                    continue
                if abs(c - xs[j][1] * sc) / max(abs(xs[j][1] * sc), 1e-9) > TOL:
                    bad += 1
        if bad:
            skip.append((cid, f"겹치는 구간에서 {bad}점이 어긋남 — 건드리지 않음"))
            continue

        tail = off + len(cv)
        add = [ym for ym, _ in col[tail:] if ym]
        if not add:
            skip.append((cid, "이미 최신"))
            continue

        newlabels = [fmt(it["labels"][-1], ym) for ym in add]
        # 계열별 마감 시점. 원자료에 정의가 바뀌는 지점이 있으면 그 앞에서 끊는다.
        # (예: KB가 2026-07 표본 개편으로 '6개광역시'를 '5개광역시'와 같은 값으로 채웠다.
        #  그대로 이으면 시장이 3.3% 떨어진 것처럼 보인다.)
        cut = conf.get("cutoff") or {}
        newseries = []
        for nm, s in zip(it["seriesNames"], it["series"]):
            xs = sheet[names.get(nm, nm)]
            lim = cut.get(nm)
            ext = []
            for k in range(len(add)):
                ym, v = xs[tail + k]
                if lim and ym > str(lim):
                    ext.append(None)
                elif v is None:
                    ext.append(None)
                else:
                    ext.append(round(v * sc, 6))
            newseries.append(list(s) + ext)
        o = ov.setdefault(it["slide"], {})
        o["labels"] = list(map(str, it["labels"])) + newlabels
        o["series"] = newseries
        o["updated"] = today
        if conf.get("source"):
            o["source"] = conf["source"]
        if conf.get("sourceUrl"):
            o["sourceUrl"] = conf["sourceUrl"]
        done.append((cid, it["title"], str(it["labels"][-1]), newlabels[-1], len(add)))

    for cid, t, f0, f1, n in done:
        print(f"  + {cid} {t[:24]:26s} {f0} → {f1}  ({n}개월 추가, {len(reg[cid].get('series') or {})}계열)")
    for cid, why in skip:
        print(f"  · {cid} {why}")

    if a.apply and done:
        json.dump(ov, open(ovp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"\n→ data/overrides.json 에 {len(done)}건 반영. 다음: python3 scripts/build.py")
    elif done:
        print("\n(미리보기입니다. 반영하려면 --apply)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
