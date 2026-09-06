#!/usr/bin/env python3
"""
사람이 건네는 표를 차트 데이터로 읽는다. set_data.py · add_chart.py 가 같이 쓴다.

받는 것:  CSV · TSV · 엑셀(.xlsx) · 붙여넣은 텍스트(탭·쉼표·세미콜론·여러 칸 공백·마크다운 |)
돌려주는 것:  {"labels": [...], "seriesNames": [...], "series": [[...], ...]}

방향은 스스로 판단한다.
  · 첫 줄이 시점(2020, 2024-05, 2025 1Q, 3월 …)으로 채워져 있으면 → 계열이 행, 시점이 열 (KOSIS·KB 엑셀 꼴)
  · 첫 열이 시점이면 → 시점이 행, 계열이 열 (보통 손으로 붙여넣는 꼴)
  · 둘 다 아니면(지역별·연령별 같은 비시계열) → 첫 열을 라벨로 본다. --transpose 로 뒤집을 수 있다.

숫자는 쉼표·%·단위 글자를 벗기고 읽는다. '-' '…' 빈칸은 결측(None)이다.
"""
import os, re, csv, io

PERIOD_RE = re.compile(
    r"^\s*(?:[’']?\d{2}|(?:19|20)\d{2})(?:\s*년)?"
    r"(?:\s*[-./년\s]?\s*(?:\d{1,2}\s*월?|[1-4]\s*[QqＱ]|[1-4]\s*분기|상반기|하반기|[Pp]))?\s*$"
    r"|^\s*\d{1,2}\s*월\s*$|^\s*(?:19|20)\d{2}\d{2}\s*$")
NUM_RE = re.compile(r"[-+]?\d[\d,]*(?:\.\d+)?|[-+]?\.\d+")


def is_period(s):
    return bool(s) and bool(PERIOD_RE.match(str(s)))


def to_num(v):
    """'1,234.5' '12.3%' '5.5조' → float,  '-' '' 'n/a' → None. 숫자가 아니면 ValueError."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("−", "-").replace("–", "-")
    if s in ("", "-", "—", "…", "...", "x", "X", "n/a", "N/A", "NA", "null", "None"):
        return None
    m = NUM_RE.search(s)
    if not m:
        raise ValueError(s)
    rest = s[:m.start()] + s[m.end():]
    if re.search(r"[A-Za-z가-힣]{3,}", rest):        # 글자가 너무 많으면 라벨이지 숫자가 아니다
        raise ValueError(s)
    return float(m.group().replace(",", ""))


def _split_line(line, sep):
    if sep == "md":
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
    elif sep == "ws":
        cells = re.split(r"\s{2,}|\t", line.strip())
    else:
        cells = next(csv.reader([line], delimiter=sep))
    return [c.strip() for c in cells]


def _guess_sep(lines):
    head = [l for l in lines if l.strip()][:5]
    if all(l.strip().startswith("|") for l in head):
        return "md"
    for sep in ("\t", ";", ","):
        cnt = [l.count(sep) for l in head]
        if cnt and min(cnt) >= 1 and max(cnt) - min(cnt) <= 1:
            return sep
    return "ws"


def rows_from_text(text):
    lines = [l for l in text.splitlines() if l.strip()]
    sep = _guess_sep(lines)
    rows = []
    for l in lines:
        if sep == "md" and re.fullmatch(r"\|?[\s:|-]+\|?", l.strip()):
            continue                                   # |---|---| 구분줄
        rows.append(_split_line(l, sep))
    return rows


def rows_from_file(path, sheet=None):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = []
        for r in ws.iter_rows(values_only=True):
            cells = []
            for c in r:
                if hasattr(c, "year"):                 # 날짜 셀 → 2024-05
                    cells.append(f"{c.year}-{c.month:02d}")
                else:
                    cells.append("" if c is None else str(c).strip())
            if any(cells):
                rows.append(cells)
        return rows
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", "replace")
    if ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        rows = [[c.strip() for c in r] for r in csv.reader(io.StringIO(text), delimiter=delim)]
        return [r for r in rows if any(r)]
    return rows_from_text(text)


def _trim(rows):
    """빈 행·열을 걷어내고 폭을 맞춘다."""
    rows = [list(r) for r in rows if any(str(c).strip() for c in r)]
    w = max(len(r) for r in rows)
    rows = [r + [""] * (w - len(r)) for r in rows]
    keep = [j for j in range(w) if any(str(r[j]).strip() for r in rows)]
    return [[r[j] for j in keep] for r in rows]


def parse(rows, transpose=None):
    """행렬 → labels / seriesNames / series. transpose=None 이면 스스로 판단한다."""
    rows = _trim(rows)
    if len(rows) < 2 or len(rows[0]) < 2:
        raise ValueError("표가 너무 작습니다 (2행 2열 이상이어야 합니다)")
    head, first_col = rows[0][1:], [r[0] for r in rows[1:]]
    head_p = sum(is_period(c) for c in head) / max(1, len(head))
    col_p = sum(is_period(c) for c in first_col) / max(1, len(first_col))
    if transpose is None:
        transpose = head_p >= 0.8 and head_p > col_p          # 시점이 가로로 누워 있다
    if transpose:
        rows = [list(x) for x in zip(*rows)]
    labels = [str(r[0]).strip() for r in rows[1:]]
    names = [str(c).strip() for c in rows[0][1:]]
    series = []
    for j in range(len(names)):
        col = []
        for r in rows[1:]:
            try:
                col.append(to_num(r[j + 1]))
            except ValueError as e:
                raise ValueError(f"숫자가 아닙니다: '{e}' (계열 '{names[j]}', 라벨 '{r[0]}')")
        series.append(col)
    return {"labels": labels, "seriesNames": names, "series": series,
            "transposed": bool(transpose)}


def load(path=None, text=None, sheet=None, transpose=None):
    rows = rows_from_file(path, sheet) if path else rows_from_text(text)
    return parse(rows, transpose)


# ── 시점 정규화: 라벨 생김새가 달라도 같은 시점이면 맞춰 본다 ─────────────
def period_key(s):
    """'2024' → ('2024',)  '2024-05' '2024.5' '24년 5월' → ('2024','05')  '2025 1Q' → ('2025','Q1')
    시점이 아니면 None."""
    if s is None:
        return None
    t = str(s).strip().replace("’", "'")
    m = re.fullmatch(r"'?(\d{2})\s*년?", t)
    if m:                                             # '24 → 2024
        return ("20" + m.group(1) if int(m.group(1)) < 70 else "19" + m.group(1),)
    m = re.fullmatch(r"((?:19|20)\d{2})\s*년?", t)
    if m:
        return (m.group(1),)
    m = re.fullmatch(r"((?:19|20)\d{2})(\d{2})", t)
    if m:                                             # 202405
        return (m.group(1), m.group(2))
    m = re.fullmatch(r"((?:19|20)\d{2})\s*[-./년]\s*(\d{1,2})\s*월?", t)
    if m:
        return (m.group(1), f"{int(m.group(2)):02d}")
    m = re.fullmatch(r"((?:19|20)\d{2})\s*[-./년]?\s*([1-4])\s*(?:[QqＱ]|분기)", t)
    if m:
        return (m.group(1), "Q" + m.group(2))
    m = re.fullmatch(r"((?:19|20)\d{2})\s*[-./년]?\s*(상반기|하반기)", t)
    if m:
        return (m.group(1), "H1" if m.group(2) == "상반기" else "H2")
    return None


def like_label(sample, key):
    """기존 라벨 생김새(sample)대로 시점 key 를 적는다."""
    s = str(sample).strip()
    y = key[0]
    sub = key[1] if len(key) > 1 else None
    if sub is None:
        if re.fullmatch(r"(19|20)\d{2}\s*년", s):
            return f"{y}년"
        if re.fullmatch(r"[’']\d{2}", s):
            return s[0] + y[2:]
        return y
    if sub.startswith("Q"):
        m = re.fullmatch(r"(19|20)\d{2}(\s*)(\d)\s*(분기|[QqＱ])", s)
        if m:
            return f"{y}{m.group(2)}{sub[1]}{m.group(4)}"
        return f"{y} {sub[1]}Q"
    if sub.startswith("H"):
        return f"{y} {'상반기' if sub == 'H1' else '하반기'}"
    m = re.fullmatch(r"(19|20)\d{2}([-./])\d{1,2}", s)
    if m:
        return f"{y}{m.group(2)}{sub}"
    if re.fullmatch(r"(19|20)\d{2}\d{2}", s):
        return f"{y}{sub}"
    if re.fullmatch(r"(19|20)\d{2}\s*년\s*\d{1,2}\s*월", s):
        return f"{y}년 {int(sub)}월"
    if re.fullmatch(r"\d{1,2}\s*월", s):
        return f"{int(sub)}월"
    if re.fullmatch(r"(19|20)\d{2}", s):                # 월간인데 연도만 반복해 찍던 차트
        return y
    return f"{y}-{sub}"
